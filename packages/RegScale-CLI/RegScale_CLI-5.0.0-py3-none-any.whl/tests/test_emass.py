#!/usr/bin/env python3
# -*- coding: utf-8 -*-
import os
import sys
import tempfile
from pathlib import Path

from openpyxl.reader.excel import load_workbook

from app.api import Api
from app.application import Application
from app.internal.login import login
from app.logz import create_logger
from app.public.emass import populate_controls, SKIP_ROWS

sys.path.append("..")  # Adds higher directory to python modules path.
TEMPLATE_WORKBOOK = ""
OUTPUT_WORKBOOK = Path()
SSP_ID = 360


class Test_Emass:
    """eMASS Test Class"""

    logger = create_logger()

    def test_init(self):
        with open("init.yaml", "r", encoding="utf-8") as file:
            data = file.read()
            self.logger.debug("init file: %s", data)
            assert len(data) > 5

    def test_create_template(self):
        """Test Catalog Code"""
        app = Application()
        api = Api(app)
        self.logger.debug(os.getenv("REGSCALE_USER"))
        self.logger.debug(os.getenv("REGSCALE_PASSWORD"))
        login(os.getenv("REGSCALE_USER"), os.getenv("REGSCALE_PASSWORD"), app=app)
        file_url = "https://regscaleblob.blob.core.windows.net/blob/eMASS_Control_Template.xlsx"
        tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx")
        f = open(tmp_file.name, "wb")
        r = api.get(url=file_url, headers={})
        f.write(r.content)
        f.close()
        self.logger.debug(f.name)
        tmp_file.close()
        global TEMPLATE_WORKBOOK
        TEMPLATE_WORKBOOK = tmp_file.name

    def test_populate_controls(self):
        """Test Profile Code"""
        app = Application()
        api = Api(app)
        self.logger.debug(os.getenv("REGSCALE_USER"))
        self.logger.debug(os.getenv("REGSCALE_PASSWORD"))
        # Get fresh token
        login(os.getenv("REGSCALE_USER"), os.getenv("REGSCALE_PASSWORD"), app=app)

        # populate the controls in the Excel workbook
        output_name = populate_controls(
            file_name=Path(TEMPLATE_WORKBOOK), ssp_id=SSP_ID, api=api
        )
        self.logger.debug(output_name.name)
        global OUTPUT_WORKBOOK
        OUTPUT_WORKBOOK = output_name

    def test_values(self):
        """Check output for populated data"""
        pass_flag: bool = False
        wb = load_workbook(OUTPUT_WORKBOOK)
        sheet = wb.active
        # iterate through rows
        for row in range(SKIP_ROWS, sheet.max_row + 1):
            compliance_status = sheet[f"L{row}"].value
            tested_by = sheet[f"N{row}"].value
            if compliance_status and tested_by:
                pass_flag = True
                self.logger.debug("Found populated data on row #:%s", row)
                break
        assert pass_flag

    def test_remove_files(self):
        os.remove(TEMPLATE_WORKBOOK)
        os.remove(OUTPUT_WORKBOOK)
