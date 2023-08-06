from datetime import timedelta
from datetime import datetime
import coursebox
# import thtools
from coursebox.thtools_base import list_dict2dict_list
# import jinjafy
import openpyxl
from coursebox.core.projects_info import populate_student_report_results
from coursebox.core.info_paths import get_paths, semester_id, semester, year, today
from coursebox.core.info_paths import core_conf
# import six
# import pybtex.database.input.bibtex
# import pybtex.plugin
# import io
# from line_profiler_pycharm import profile
import time

# @profile
def xlsx_to_dicts(xlsx_file,sheet=None, as_dict_list=False):
    # print("Loading...", xlsx_file, sheet, as_dict_list)
    t0 = time.time()
    wb = openpyxl.load_workbook(xlsx_file, data_only=True, read_only=True)
    if not sheet:
        ws = wb.worksheets[0]
    else:
        ws = [ws for ws in wb.worksheets if ws.title == sheet]
        if len(ws) < 1:
            return None
        else:
            ws = ws.pop()
    # print(time.time()-t0)
    # dd = []
    # key_cols = [j for j in range(ws.max_column) if ws.cell(row=1, column=j + 1).value is not None]
    # print(time.time()-t0, ws.max_row)
    # np.array([[i.value for i in j[1:5]] for j in ws.rows])

    import numpy as np
    A = np.array([[i.value for i in j] for j in ws.rows])
    # print(time.time() - t0, ws.max_row, len(key_cols))


    # for j in range(A.shape[1]):




    a = 234

    # for i in range(1, ws.max_row):
    #     rdict = {}
    #     if not any( [ws.cell(row=i+1, column=j+1).value is not None for j in key_cols] ):
    #         continue
    #     for j in key_cols:
    #         key = ws.cell(row=1, column=j+1).value
    #         if key is not None:
    #             key = key.strip() if isinstance(key,str) else key
    #             value = ws.cell(row=i + 1, column=j + 1).value
    #             value = value.strip() if isinstance(value,str) else value
    #             if isinstance(value, str):
    #                 if value == 'True':
    #                     value = True
    #                 if value == 'False':
    #                     value = False
    #             rdict[key] = value
    #     dd.append(rdict)

    # print(time.time()-t0)

    A = A[:, A[0] != None]
    A = A[(A != None).sum(axis=1) > 0, :]

    dd2 = []
    for i in range(1, A.shape[0]):
        A[A == 'True'] = True
        A[A == 'False'] = False

        d = dict(zip(A[0, :].tolist(), [a.strip() if isinstance(a,str) else a for a in A[i, :].tolist() ]))
        dd2.append(d)

    # print(time.time() - t0)
    dd = dd2
    # if dd != dd2:
    #     for k in range(len(dd)):
    #         if dd[k] != dd2[k]:
    #             print(k)
    #             print(dd)
    #             print(dd2)
    #     assert False
    #     print("BAd!")
    if as_dict_list:
        dl = list_dict2dict_list(dd)
        for k in dl.keys():
            x = [v for v in dl[k].tolist() if v is not None]
            if len(x) == 1: x = x.pop()
            dl[k] = x
        dd = dl
    wb.close()
    # print("xlsx2dicts", time.time()-t0)
    return dd

def get_enrolled_students():
    paths = get_paths()
    students = xlsx_to_dicts(paths['information.xlsx'], sheet='students')
    students2 = {}
    for s in students:
        s2 = {}
        if s['Study number']:
            s['Username'] = s['Study number']
        for k in s.keys():
            k2 = k.lower().replace(" ", "")
            if k2 == "studynumber": continue
            if k2 == "username":
                k2 = "id"
                if not s[k] or len(s[k]) == 0:
                    print("Bad study id: ")
                    print(s)
                    raise Exception("malformed course configuration file, bad student id")
            s2[k2] = s[k]

        id = s2['id']
        students2[id] = s2
    return students2

def get_instructors():
    paths = get_paths()
    instructors = xlsx_to_dicts(paths['information.xlsx'], sheet='instructors')
    return instructors

def continuing_education():
    return coursebox.core.info_paths.core_conf['continuing_education_mode']

def first_day_of_class(info):
    if continuing_education():
        first_day_of_class = datetime(year=year(), month=info['first-month'], day=info['first-day'], hour=info['hour'][0], minute=info['minute'][0])
    else:
        mo_first = datetime(year=year(), month=1 if semester() == 'spring' else 8, day=1, hour=13, minute=0)
        # scroll to monday
        while mo_first.weekday() != 0: #strftime('%A') is not 'Monday':
            mo_first += timedelta(days=1)
        # add 4 weeks to get into 13 week period
        for _ in range(4):
            mo_first += timedelta(days=7)

        dow = int(info.get('day_of_week', 1))
        while mo_first.weekday() != dow:
            mo_first += timedelta(days=1)

        first_day_of_class = mo_first
    return first_day_of_class


def lectures(info, pensum=None):
    ow = timedelta(days=7)
    d = first_day_of_class(info)

    holiday = int(info['holiday_week']) if "holiday_week" in info else (9 if semester() == "spring" else 6)
    paths = get_paths()
    lectures = []
    lecture_info = xlsx_to_dicts(paths['information.xlsx'], sheet='lectures')

    for lecture in lecture_info:
        em = lecture["resources"]
        r = []
        if em:
            ems = em.split("\n")
            for e in ems:
                e = e.strip()
                url = e[:e.find(" ")]
                description = e[e.find(" ") + 1:]
                shorturl = url[:url.find("/",url.find("."))]
                r.append( {'url': url, 'shorturl': shorturl, 'description': description})
        lecture["resources"] = r
        if pensum is not None:

            rd, html = lecture['reading'], ""
            if rd is None:
                rd = ""

            while True:
                i = rd.find("\\cite")
                if i < 0: break
                j = rd.find("}", i)
                html += rd[:i]
                cite = rd[i:j + 1]
                rd = rd[j+1:]
                if cite.find("[") > 0:
                    sr = cite[cite.find("[")+1:cite.find("]")]
                else:
                    sr = None
                key = cite[cite.find("{")+1:cite.find("}")]
                html += "[<b>" + pensum[key]['label'] + "</b>" + (", " + sr if sr is not None else "") + "]"
                pensum[key]['suggested'] = True
            html += rd
            lecture['reading_html'] = html

    if continuing_education():
        ice = xlsx_to_dicts(paths['information.xlsx'], sheet='ce', as_dict_list=True)
        holiday = -1
        dd = [ice['day'][i] - ice['day'][i-1] for i in range(1, len(ice['day']))]
        dd.append(0)

    for i in range(0, len(lecture_info)):
        l = dict()

        l['year'] = d.year
        l['month'] = d.strftime('%B')
        l['day'] = d.day
        l['date'] = d
        l['preceded_by_holiday'] = i == holiday
        l = {**l, **date2format(d)}
        # l['date_short'], l['date_long'] = date2format(d)
        # l['date_long'] =

        if not continuing_education():
            ehw = int(info.get('extraordinary_holiday_week', -1))
            if 'extraordinary_holiday_week' in info and int(lecture_info[i]['number']) == int(info['extraordinary_holiday_week']) and ehw == 13:

                dow = int(info['day_of_week'])
                if dow == 4:  # friday
                    d2 = d + timedelta(days=4)  #
                    l['month'] = d2.strftime('%B')
                    l['day'] = d2.day
                    l['date'] = d2
                    l['extraordinary'] = True

            d = d + ow


            if i == holiday - 1: d = d + ow
            if d.month == 5 and d.day == 8: # grundlovsdag
                d += timedelta(days=4)


        else:
            d = d + timedelta(days=dd[i-0] if i > 1 else 0)
            d = d.replace(hour=ice['hour'][i-1], minute=ice['minute'][i-1])



        linfo = lecture_info[i]

        ir = linfo.get('reading', "")
        linfo['reading_long'] = ir.replace("C", "Chapter ") if ir else ""

        hwp =  linfo['homework_problems']
        linfo['homework_problems_long'] = hwp.replace("P", "Problem ") if hwp else ""
        if linfo["learning_objectives"]:
            linfo["learning_objectives"] = [s.strip() for s in linfo["learning_objectives"].split("\n")]
        l.update(linfo)
        lectures.append(l)
    return lectures, pensum
def date2format(nd):
    ab = 'st'
    if nd.day == 2:
        ab = "nd"
    elif nd.day == 3:
        ab = 'rd'
    elif nd.day >= 4:
        ab = 'th'

    latex_long = f"{nd.strftime('%A')} {nd.day}{ab} {nd.strftime('%B')}, {nd.year}"
    latex_short = f"{nd.strftime('%B')} {nd.day}{ab}, {nd.year}"
    return {'latex_short': latex_short,
            'latex_long': latex_long,
            'latex_abbrev': f"{nd.strftime('%b')} {nd.day}{ab}",
            'latex_abbrev_year': f"{nd.strftime('%b')} {nd.day}{ab}, {nd.year}",
            }


    # return latex_short, latex_long

def get_forum(paths):
    a = xlsx_to_dicts(paths['information.xlsx'], sheet='forum', as_dict_list=True)
    if a is None:
        return a
    from collections import defaultdict
    dd = defaultdict(lambda: [])
    kk = list(a.keys())[0]
    for i, k in enumerate(kk.split(",")):
        k = k.replace("[", "")
        k = k.replace("]", "")
        k = k.split(" ")[0]
        for v in a[kk]:
            dd[k.lower()].append(v.split(",")[i])

    n = len(list(dd.values())[0])
    d2 = []
    for i in range(n):
        d2.append({k: v[i] for k, v in dd.items()})
    return d2

# @profile
def class_information(verbose=False):
    course_number = core_conf['course_number']
    piazza = 'https://piazza.com/dtu.dk/%s%s/%s' % (semester().lower(), year(), course_number)
    paths = get_paths()
    teachers = xlsx_to_dicts(paths['information.xlsx'], sheet='teachers')
    students, all_groups = populate_student_report_results( get_enrolled_students(), verbose=verbose)
    continuing_education_mode = core_conf['continuing_education_mode']
    faq = xlsx_to_dicts(paths['information.xlsx'], sheet='faq')

    d = {'year': year(),
         'piazza': piazza, # deprecated.
         'course_number': course_number,
         'semester': semester(),
         # 'reports_handout': [1,6], # Set in excel conf.
         # 'reports_handin': [6, 11], # set in excel conf.
         'semester_id': semester_id(),
         'today': today(),
         'instructors': get_instructors(),
         'students': students,
         'teachers': teachers,
         "CE": continuing_education_mode,
         "all_groups": all_groups,
         "faq": faq,
         'forum': get_forum(paths),
         }

    written_exam = xlsx_to_dicts(paths['information.xlsx'], sheet='exam', as_dict_list=True)
    if "solution_q" in written_exam:
        written_exam['solution'] = {n:a for n,a in zip( written_exam['solution_q'], written_exam['solution_a'] ) }

    d['written_exam'] = written_exam

    gi = xlsx_to_dicts(paths['information.xlsx'], sheet='general_information', as_dict_list=True)
    for (k, v) in zip(gi['key'], gi['value']):
        if v == 'True':
            v = True
        if v == 'False':
            v= False
        gi[k] = v
    del gi['key']
    del gi['value']

    from snipper.load_citations import get_bibtex, get_aux
    if "pensum_bib" in gi:
        bibtex = get_bibtex(paths['02450public'] + "/" + gi['pensum_bib'])
        # refs, nrefs = get_references(paths['02450public'] + "/" + gi['pensum_bib'], gi)
        # d['references'], d['new_references'] = refs, nrefs
        cmds = []
        ls = lambda x: x if isinstance(x, list) else [x]
        if 'tex_command' in gi:
            for cmd, aux, display in zip(ls(gi['tex_command']), ls(gi['tex_aux']), ls(gi['tex_display'])):
                cm = dict(command=cmd, aux=get_aux(paths['02450public'] + "/"+aux), output=display)
                cmds.append(cm)
        d['references'] = dict(bibtex=bibtex, commands=cmds)


    else:
        print("[info]", "No bibtex rereferences specified. Check configuration file. ")
        d['references'] = dict(commands=[], bibtex={}) #['bibtex'] = None
    d.update(gi)
    # set first day of class if CE
    if continuing_education_mode:
        ice = xlsx_to_dicts(paths['information.xlsx'], sheet='ce', as_dict_list=True)
        d.update(ice)

    d['lectures'], d['references']['bibtex'] = lectures(info=d, pensum=d['references']['bibtex'])

    d['first_day_of_class'] = first_day_of_class(info=d)
    d['day_of_week_str'] = d['first_day_of_class'].strftime('%A')
    if "piazza" in gi:
        d['piazza'] = gi['piazza']

    for k in ['freeze_report_evaluation', 'freeze_grades']:
        freeze = gi[k]
        freeze = freeze == "True" if isinstance(freeze, str) else freeze
        freeze = freeze[0] if isinstance(freeze, list) else freeze
        gi[k] = freeze

    for k,v in core_conf.items():
        d[k] = v

    d['CE2'] = gi.get("days", 5) == 2 if continuing_education_mode else False
    d['CE5'] = gi.get("days", 5) == 5 if continuing_education_mode else False

    d['freeze_report_evaluation'] = d['freeze_report_evaluation'] == 'True'
    d['freeze_grades'] = d['freeze_grades'] == 'True'

    d['rooms'] = xlsx_to_dicts(paths['information.xlsx'], sheet='rooms')
    fix_instructor_comma(d['rooms'], d['instructors'])

    d['teams'] = xlsx_to_dicts(paths['information.xlsx'], sheet='teams')
    fix_instructor_comma(d['teams'], d['instructors'])

    if 'reports_delta' in d:
        print(234)



    if 'handin_day_delta' in d:
        d['reports_info'] = {}
        for k, r in enumerate(d['reports_handin']):
            ri = {}
            d['reports_info'][r] = ri
            # print(d['lectures'][r-1]['date'], r)
            nd = d['lectures'][r-1]['date'] + timedelta(days=int(d['handin_day_delta']))
            ri['date'] = nd
            ri['html'] = f"{nd.day} {nd.strftime('%b')}"
            # ab = 'st'
            # if nd.day == 2:
            #     ab = "nd"
            # elif nd.day == 3:
            #     ab = 'rd'
            # elif nd.day >= 4:
            #     ab = 'th'
            # latex_short, latex_long = date2format(nd)

            # ri['latex_long'] = latex_long # f"{nd.strftime('%A')} {nd.day}{ab} {nd.strftime('%B')}, {nd.year}"
            # ri['latex_short'] = latex_short # f"{nd.strftime('%B')} {nd.day}{ab}, {nd.year}"
            ri = {**ri, **date2format(nd)}



            d['reports_info'][k] = ri


    ppi = core_conf.get('post_process_info', None)
    if ppi is not None:
        d = ppi(paths, d)

    r_details = {}

    def get_lecture_date(lecture_id, delta_days=0):
        ri = {}
        ri['lecture'] = lecture_id

        if lecture_id is None:
            return ri
        l = [l for l in d['lectures'] if l['number'] == lecture_id][0]

        nd = l['date'] + timedelta(days=delta_days)

        ri['date'] = nd
        ri['html'] = f"{nd.day} {nd.strftime('%b')}"
        # ab = 'st'
        # if nd.day == 2:
        #     ab = "nd"
        # elif nd.day == 3:
        #     ab = 'rd'
        # elif nd.day >= 4:
        #     ab = 'th'
        # ri['latex_long'] = f"{nd.strftime('%A')} {nd.day}{ab} {nd.strftime('%B')}, {nd.year}"
        # ri['latex_short'] = f"{nd.strftime('%B')} {nd.day}{ab}, {nd.year}"
        ri = {**ri, **date2format(nd)}
        # d['reports_info'][k] = ri

        return ri

    # TH: This is the new way of specifying projects. Change to this datastructure gradually.
    reports = xlsx_to_dicts(paths['information.xlsx'], sheet='reports', as_dict_list=False)
    if reports is not None:
        d['reports'] = {}
        for r in reports:
            d['reports'][r['id']] = r
            r['handin'] = get_lecture_date(r['handin'], delta_days=int(d['handin_day_delta']))
            r['handout'] = get_lecture_date(r['handout'], delta_days=0)
            r['exercises'] = [e.strip() for e in r['exercises'].split(",") if len(e.strip()) > 0]


    ice = xlsx_to_dicts(paths['information.xlsx'], sheet='ce', as_dict_list=True)
    return d

def fix_instructor_comma(dd, instructors):
    for r in dd:
        ri_shortnames = [i.strip().lower() for i in r['instructors'].split(",")]
        ri = []
        for sn in ri_shortnames:
            di = [i for i in instructors if i['shortname'] == sn ]
            if not di:
                print("Did not find shortname: " + sn + ". This seems bad.")
            ri += di
        r['instructors'] = ri