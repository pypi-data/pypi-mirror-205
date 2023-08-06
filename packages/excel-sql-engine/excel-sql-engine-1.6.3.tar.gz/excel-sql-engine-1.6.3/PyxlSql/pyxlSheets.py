
# ---------------------------------------------------------------------------------------------------------------
# PyxlSQL project
# This program and library is licenced under the European Union Public Licence v1.2 (see LICENCE)
# developed by fabien.battini@gmail.com
# ---------------------------------------------------------------------------------------------------------------

import re
import inspect
import openpyxl
import openpyxl.styles
from PyxlSql.pyxlErrors import PyxlSqlColumnError
from copy import copy

class PyxlStyle:
    """Style for columns, limited features"""

    def __init__(self,cell):
        self.number_format = cell.number_format
        self.style = cell.style
        self.fill = copy(cell.fill)
        self.font = copy(cell.font)
        self.alignment = copy(cell.alignment)

    def copy_to_cell(self,cell):
        # for some strange reason, the order of the following statements is important.
        # probably setting the style resets the number format
        cell.style = self.style
        cell.fill = copy(self.fill)
        cell.font = copy(self.font)
        cell.number_format = self.number_format
        cell.alignment = copy(self.alignment)


# ---------------------------------------------------------------------------------------------------------------
# class NamedWS
# ---------------------------------------------------------------------------------------------------------------


class PyxlSheet:
    """
    Additional metadata for each named sheet in the NamedWB
    NB: Column names are *case-sensitive* (NOT converted to lower case)
    """
    general_style = 'General'
    int_style = '#,##0'
    percent_style = '0.0%'
    euro_style = '_ * #,##0_) [$€-x-euro1]_ '
    euro2_style = '_ * #,##0.00_) [$€-x-euro1]_ '

    def __init__(self, workbook, sheet, to_read=True,
                 first_row=1, first_column=1, last_row=None, last_column=None, font=None):
        sheet.auto_filter.ref = None    # to avoid issues when auto_filter is not correctly managed by OpenPyxl

        self.openpyxl_sheet = sheet
        self.father = workbook

        self.full_name = self.father.filename + ':[' + sheet.title + ']'
        self.current_row = sheet.max_row + 1    # first free row
        self.first_row = first_row or 1
                # number of the row where the column name is stored
        self.last_row = last_row
        self.first_column = first_column or 1
        self.last_column = last_column or sheet.max_column+1
        self.columns = []                       # a list of columns
        self.column_names = {}                  # a dictionary: string --> Column number
        self.column_styles = {}                 # a dictionary: string --> style
        self.indexes = {}                       # a dictionary: index -->
        self.book_default_font = openpyxl.styles.Font(name='Century Gothic', size=11) if font is None else font
        self.data_is_read = False
        if to_read:
            self.read_data(first_row, first_column, last_row, last_column, font)



    def read_data(self,  first_row=1, first_column=1, last_row=None, last_column=None, font=None):
        if self.data_is_read:
            return
        self.current_row = self.openpyxl_sheet.max_row + 1  # first free row
        self.first_row = first_row or 1
        # number of the row where the column name is stored
        self.last_row = last_row
        self.first_column = first_column or 1
        self.last_column = last_column or self.openpyxl_sheet.max_column + 1
        self.columns = []                    # a list of columns
        self.column_names = {}               # a dictionary: string --> Column number
        self.column_styles = {}              # a dictionary: string --> style
        self.indexes = {}                    # a dictionary: index -->
        self.book_default_font = openpyxl.styles.Font(name='Century Gothic', size=11) if font is None else font

        for i in range(self.first_column, self.last_column):
            col_name = self.openpyxl_sheet.cell(row=first_row, column=i).value
            if col_name is not None and col_name != "" and col_name != " ":
                col_name = str(col_name)  # str(), because could be an integer, typically a year
                self.column_names[col_name] = i
                self.columns.append(col_name)
                cell = self.openpyxl_sheet.cell(row=first_row + 1, column=i)
                if cell is not None:
                    self.column_styles[col_name] = PyxlStyle(cell)
        self.data_is_read = True
    def get_name(self):
        return self.openpyxl_sheet.title

    def rename(self, new_name:str):
        self.openpyxl_sheet.title = new_name
        self.full_name = self.father.filename + ':[' + new_name + ']'

    def get_row_range(self):
        """returns the range of ACTIVE rows"""
        return range(self.first_row + 1, self.current_row)

    def get_start_of_range(self):
        return self.first_row + 1

    def find_column(self, column_name):
        """returns the column number from its name"""
        if column_name is None:
            raise PyxlSqlColumnError("find_column(None)", self.full_name)

        lc_column_name = str(column_name)  # .lower()
        if lc_column_name not in self.column_names:
            raise PyxlSqlColumnError(column_name, self.full_name)
        return self.column_names[lc_column_name]

    def get_column_range(self):
        return range(self.first_column, self.last_column)

    @staticmethod
    def build_key(*args):
        """
        Creates the hash key for a list of descriptors,
        Typically (Column name, Row name)

        CAVEAT: all identifiers are NO MORE turned into lower case
        """
        key = ""
        for v in args:
            key += "" if v is None else (":" + str(v))  # str(v).lower())
        return key

    def get_cell(self, row_nb, column_name):
        col = self.find_column(column_name)
        if row_nb is None:
            raise PyxlSqlColumnError("Undefined",
                                     f"source:{str(inspect.currentframe().f_lineno)}" +
                                     f":{str(inspect.currentframe().f_code.co_filename)}")
        return self.openpyxl_sheet.cell(row=row_nb, column=col)

    def get_val_by_nb(self, row_nb, col):
        cell = self.openpyxl_sheet.cell(row=row_nb, column=col)
        return None if cell.value is None or cell.value == "" else cell.value

    def get_val(self, row_nb, column_name):
        cell = self.get_cell(row_nb, column_name)
        return None if cell.value is None or cell.value == "" else cell.value

    def get_string(self, row_nb, column_name):
        cell = self.get_cell(row_nb, column_name)
        if cell is None or cell.value is None or cell.value == "None":
            return ""
        return str(cell.value)

    @staticmethod
    def get_html_value_from_cell(cell):

        if cell is None or cell.value is None or cell.value == "None":
            return ""
        if isinstance(cell.value, str) and (cell.value != '' and cell.value[0] == '='):
            # this is a formula
            return cell.value

        if cell.is_date:
            proto = cell.number_format
            val = proto.replace("mm-dd-yy", "{:02d}/{:02d}/{:04d}".format(cell.value.month, cell.value.day,
                                                                          cell.value.year))  # strange enough
            val = val.replace("yyyy", "{:4d}".format(cell.value.year))
            val = val.replace("h:mm", "{:d}:{:02d}".format(cell.value.hour, cell.value.minute))
            val = val.replace("yy", "{:02d}".format(cell.value.year % 100))
            val = val.replace("y", "{:d}".format(cell.value.year))
            val = val.replace("mm", "{:02d}".format(cell.value.month))
            val = val.replace("m", "{:d}".format(cell.value.month))
            val = val.replace("dd", "{:02d}".format(cell.value.day))
            val = val.replace("d", "{:d}".format(cell.value.day))
            val = val.replace("h", "{:d}".format(cell.value.hour))
            val = val.replace("s", "{:02d}".format(cell.value.second))
            val = val.replace(";@", '')
            val = val.replace("\\ ", '&nbsp;')
            val = val.replace("\\-", '-')
            return val

        number_formats = {
            "0.": ":.0f",
            "0.0": ":.1f",
            "0.00": ":.2f",
            "0.000": ":.3f",
            "0.0000": ":.4f",
            "0.00000": ":.5f",
            "0.000000": ":.6f",
            "0.0000000": ":.7f",
            "0.00000000": ":.8f",
            "0.000000000": ":.9f",
            "0.0000000000": ":.10f",
            "0.00000000000": ":.11f",
            "0.000000000000": ":.12f",
            "0%": ":.0f",
            "0.0%": ":.1f",
            "0.00%": ":.2f",
            "0.000%": ":.3f",
            "0.0000%": ":.4f",
            "0.00000%": ":.5f",
            "0.000000%": ":.6f",
            "0.0000000%": ":.7f",
            "0.00000000%": ":.8f",
            "0.000000000%": ":.9f",
            "0.0000000000%": ":.10f",
        }

        if cell.number_format in number_formats.keys():
            proto = "{" + number_formats[cell.number_format] + '}'
            if cell.number_format[-1] == '%':
                try:
                    val = proto.format(float(cell.value) * 100)
                    return val + "%"
                except ValueError:
                    # sometimes, the format of the cell is inherited from a previous value,
                    # and therefore is set as %, while the content is indeed a string!
                    return str(cell.value)

            val = proto.format(cell.value)
            return val
        return str(cell.value)

    def get_html_value(self, row_nb, column_name):
        cell = self.get_cell(row_nb, column_name)
        return self.get_html_value_from_cell(cell)


    def get_cell_color(self, cell):
        return self.father.get_rgb_color(cell.fill.fgColor)



    def get_html_style(self, row_nb, column_name):
        cell = self.get_cell(row_nb, column_name)
        return self.father.get_html_style_from_cell(cell)

    def get_float(self, row_nb, column_name):
        cell = self.get_cell(row_nb, column_name)
        if cell is None:
            return 0
        val = cell.value

        if val is None or val == "":
            return 0

        if isinstance(val, str) and val.startswith("="):
            val = val[1:]  # remove leading =
            int_val = eval(val)  # a tentative to manage simple cells, such as '=12.56-24.9'
            # print("EVAL", val, intVal)
            # TODO: Track Errors
            return int_val

        return float(val)  # ********** CAVEAT! must be Float, because can be less than 1

    def get_style(self, column_name)->PyxlStyle :
         lc_column_name = str(column_name)  # .lower()
         if lc_column_name not in self.column_names:
             raise PyxlSqlColumnError(column_name, self.full_name)
         return self.column_styles[lc_column_name]

    # def get_index(self, *column_names):
    #     """
    #     :param self: a sheet
    #     :param column_names: the list of columns to be indexed
    #     :return: the index, i.e. the dictionary (value of the indexed columns) --> row number
    #
    #     if the index was not yet created, creates it
    #     """
    #     key = NamedWS.build_key(*column_names)  # Manages SEVERAL columns
    #     if key not in self.indexes:
    #         index_hash = {}
    #         for row in self.get_row_range():
    #             values = []
    #             for col in column_names:
    #                 values.append(self.get_val(row, col))
    #             val_key = NamedWS.build_key(*values)
    #             if val_key not in index_hash:
    #                 index_hash[val_key] = []
    #             index_hash[val_key].append(row)
    #         self.indexes[key] = index_hash
    #     return self.indexes[key]

    # ------------ set methods

    def set_value(self, row_nb: int, column_name: str, value, number_format=None) -> None:
        cell = self.get_cell(row_nb, column_name)
        assert cell is not None, "INTERNAL: set_value"

        if isinstance(value, str):
            value = re.sub("^ *", "", value)
            value = re.sub(" *$", "", value)

        cell.value = value
        cell.font = self.book_default_font
        style = number_format if number_format is not None else self.get_style(column_name)
        if style is not None:
            style.copy_to_cell(cell)

        self.current_row = max(self.current_row, row_nb+1)  # current_row is the first FREE cell

    # UNUSED, for future PIVOT
    # def create_row(self, init):
    #     for column, value in init.items():
    #         self.set_value(self.current_row, column, value)
    #
    #     self.current_row += 1
    #     return self.current_row - 1

    @staticmethod
    def get_col_from_int(column_nb):
        column_nb -= 1
        my_max = ord('Z') - ord('A') + 1
        low = column_nb % my_max
        res = chr(ord('A') + low)
        high = int((column_nb - low) / my_max)

        if high > 0:
            res = chr(ord('A')+high-1) + res
        return res
    def get_reference(self, row_nb, column_name, absolute=True, same_sheet=True):
        if same_sheet:
            sheet_ref = ""
        else:
            sheet_ref = "'" + self.openpyxl_sheet.title + "'!"

        if absolute:
            ref = "{}${}${}".format(sheet_ref, self.get_col_from_int(self.find_column(column_name)),
                                    row_nb)
        else:
            ref = "{}{}{}".format(sheet_ref, self.get_col_from_int(self.find_column(column_name)),
                                  row_nb)
        return ref