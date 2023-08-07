
# ---------------------------------------------------------------------------------------------------------------
# PyxlSQL project
# This program and library is licenced under the European Union Public Licence v1.2 (see LICENCE)
# developed by fabien.battini@gmail.com
# ---------------------------------------------------------------------------------------------------------------

import re
import os
import importlib
import openpyxl
import openpyxl.styles
from PyxlSql.pyxlErrors import PyxlSqlSheetError, PyxlSqlError, PyxlSqlExecutionError, PyxlSqlInternalError
from PyxlSql.pyxlSheets import PyxlSheet

# ---------------------------------------------------------------------------------------------------
# class NamedWB
# ---------------------------------------------------------------------------------------------------


class PyxlWorkbook:
    """
    class to manage excel workbooks structured as a database
    """
    def __init__(self, file_name, create=False, first_row=1, first_column= 1, last_row=None, last_column=None,
                 font=None):


        if file_name is None:
            raise PyxlSqlError(f"Call PyxlWorkbook with file '{None}'", "command line arguments")

        self.filename = file_name
        self.file_path= os.path.dirname(self.filename) or '.'

        try:
            self.wb: openpyxl.workbook = openpyxl.load_workbook(filename=self.filename)
        except OSError as error:
            print(f"FATAL ERROR: Cannot open '{file_name}' : {str(error)}, aborting")
            exit(-1)

        self.sheets = {}  # a dictionary  string --> NamedWS

        self.book_default_font = openpyxl.styles.Font(name='Century Gothic', size=11) if font is None else font
        self.wbs = {}       # all workbooks referenced by this one
        self.imported = {}  # dictionary of imported symbols, will be used when eval is called
        self.sheets_to_delete = []
        self.import_module("functools")
        self.theme = self.init_theme()
        self.first_row = first_row
        self.last_row = last_row
        self.first_column = first_column
        self.last_column = last_column
        if create:
            self.wb = openpyxl.Workbook()
            return


    # <a:dk1><a:sysClr val="windowText" lastClr="000000"/></a:dk1>
    # <a:lt1><a:sysClr val="window" lastClr="FFFFFF"/></a:lt1>
    # <a:dk2><a:srgbClr val="44546A"/></a:dk2>                   # noqa
    # <a:lt2><a:srgbClr val="E7E6E6"/></a:lt2>                   # noqa
    # <a:accent1><a:srgbClr val="5B9BD5"/></a:accent1>           # noqa
    # <a:accent2><a:srgbClr val="ED7D31"/></a:accent2>           # noqa
    # <a:accent3><a:srgbClr val="A5A5A5"/></a:accent3>           # noqa
    # <a:accent4><a:srgbClr val="FFC000"/></a:accent4>           # noqa
    # <a:accent5><a:srgbClr val="4472C4"/></a:accent5>           # noqa
    # <a:accent6><a:srgbClr val="70AD47"/></a:accent6>           # noqa
    # <a:hlink><a:srgbClr val="0563C1"/></a:hlink>               # noqa
    # <a:folHlink><a:srgbClr val="954F72"/></a:folHlink>         # noqa
    #
    def init_theme(self):
        theme = []
        theme_desc = self.wb.loaded_theme.decode('utf-8')

        m = re.search(r'<a:lt1><a:sysClr val=".*" lastClr="([\dABCDEF]*)"/></a:lt1>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")
        # CAVEAT: for some strange reason, the first 4 items do not appear in the appropriate order inside the xml
        # we have to reorder to get the Windows file color scheme !

        m = re.search(r'<a:dk1><a:sysClr val=".*" lastClr="([\dABCDEF]*)"/></a:dk1>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:lt2><a:srgbClr val="([\dABCDEF]*)"/></a:lt2>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:dk2><a:srgbClr val="([\dABCDEF]*)"/></a:dk2>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")


        m = re.search(r'<a:accent1><a:srgbClr val="([\dABCDEF]*)"/></a:accent1>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:accent2><a:srgbClr val="([\dABCDEF]*)"/></a:accent2>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:accent3><a:srgbClr val="([\dABCDEF]*)"/></a:accent3>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:accent4><a:srgbClr val="([\dABCDEF]*)"/></a:accent4>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:accent5><a:srgbClr val="([\dABCDEF]*)"/></a:accent5>', theme_desc)
        if m:
            theme.append(m.group(1))
        else:
            raise PyxlSqlInternalError("Excel wrong file format for color theme")

        m = re.search(r'<a:accent6><a:srgbClr val="([\dABCDEF]*)"/></a:accent6>', theme_desc)
        if m:
            theme.append(m.group(1))

        m = re.search(r'<a:folHlink><a:srgbClr val="([\dABCDEF]*)"/></a:folHlink>', theme_desc)
        if m:
            theme.append(m.group(1))

        return theme

    def get_sheet(self, sheet_name, first_row=None, first_column=None, last_row=None, last_column=None,
                  raise_exception=True, to_read=True):
        first_row = first_row or self.first_row

        if sheet_name is None:
            return None
        if sheet_name in self.sheets:
            return self.sheets[sheet_name]

        if sheet_name in self.wb:
            sheet = self.wb[sheet_name]
            self.sheets[sheet.title] = PyxlSheet(self, sheet, to_read=to_read,
                                                 first_row=first_row,last_row=last_row,
                                                 first_column=first_column,last_column=last_column)
            return self.sheets[sheet_name]

        m = re.search("\"?\'?([^\"\'[]+)\"?\'?(\\[(.*)])?", sheet_name)
        if m is not None:
            first_name: str = m.group(1)
            second_name: str = m.group(3)
            if first_name is not None and second_name is not None:
                workbook = self.get_workbook(first_name)
                if workbook is not None:
                    if second_name in workbook.sheets:
                        return workbook.sheets[second_name]
                    if second_name in workbook.wb.sheetnames:
                        sheet = workbook.wb[second_name]

                        workbook.sheets[sheet.title] = PyxlSheet(self, sheet, to_read=to_read,
                                                                 first_row=first_row,last_row=last_row,
                                                                 first_column=first_column,last_column=last_column)
                        return workbook.sheets[second_name]

            # here, we have not found the sheet

        if raise_exception:
            raise PyxlSqlSheetError(sheet_name, "workbook")
        return None

    def delete_sheet(self, sheet_name):
        """Deletes the sheet from the workbook"""
        sheet = self.get_sheet(sheet_name)
        if sheet is not None:
            del self.wb[sheet_name]

    def rename_sheet(self, old_sheet_name:str, new_sheet_name:str):
        """Rename the sheet from the workbook"""
        sheet = self.get_sheet(old_sheet_name)
        # sheet cannot be None, get_sheet would have raised an exception

        sheet.rename(new_sheet_name)
        self.sheets[new_sheet_name]=sheet
        self.sheets.pop(old_sheet_name)

    def save(self, file_name):
        current_dir = os.path.dirname(os.path.realpath(self.filename))
        try:
            self.wb.save(current_dir + "/" + file_name)
        except OSError as err:
            raise PyxlSqlExecutionError(f" file '{current_dir}\\{file_name}' is read-only", str(err))

    @staticmethod
    def local_open(file_name: str, mode: str):
        try:
            ins = open(file_name, mode)
        except OSError as error:
            raise PyxlSqlExecutionError(f"FATAL ERROR: Cannot open('{file_name}',{mode})", str(error))
        return ins

    @staticmethod
    def norm(f:float)->int:
        i = int(f)
        i = max(0,i)
        i = min(255,i)
        return i

    @staticmethod
    def hex2(i:int):
        if i < 16:
            return "0"+hex(i)[2:]
        return hex(i)[2:]
    @staticmethod
    def hsl_to_rgb(t,s,l):
        lp = l/255
        c = (1 - abs(2*lp -1))*s
        tp = t*6/256
        m = l - c/2
        x = c*(1-abs(tp % 2 -1))
        # if t == 0 and False:
        #     rp,gp,bp=0,0,0
        if 0 <= tp < 1:
            rp,gp,bp = c,x,0
        elif 1 <= tp < 2:
            rp,gp,bp = x,c,0
        elif 2 <= tp < 3:
            rp,gp,bp = 0,c,x
        elif 3 <= tp < 4:
            rp,gp,bp = 0,x,c
        elif 4 <= tp < 5:
            rp,gp,bp = x,0,c
        else:
            rp,gp,bp = c,0,x
        r = PyxlWorkbook.norm(rp + m)
        g = PyxlWorkbook.norm(gp + m)
        b = PyxlWorkbook.norm(bp + m)
        return PyxlWorkbook.hex2(r) + PyxlWorkbook.hex2(g) + PyxlWorkbook.hex2(b)
    @staticmethod
    def rgb_to_hsl(rgb):
        # see https://fr.wikipedia.org/wiki/Teinte_saturation_luminosité
        r = int(rgb[0:2],16)
        g = int(rgb[2:4],16)
        b = int(rgb[4:6],16)
        ma = max(r,g,b)
        mi = min(r,g,b)
        c = ma - mi

        if c == 0:
            tp = 0
        elif ma == r:
            tp = ((g - b) / c) % 6
        elif ma == g:
            tp = ((b - r) / c + 2) % 6
        else:
            tp = ((r - g) / c + 4) % 6


        t =  (tp * 255)/6 # if rgb != "FFFFFF" else 170
        l =  ((ma+mi)/2)
        lp = l / 255
        s = 0.0 if  lp == 0 or lp == 1 else  (c / (1 - abs(2 * lp - 1)))
        return t,s,l

    @staticmethod
    def tint_rgb(rgb, tint) -> str:
        t, s, l = PyxlWorkbook.rgb_to_hsl(rgb)
        if l == 255 or (t == 0 and l != 0):      #   strange behavior, but tests leads to this !
            l = l * (1+tint)
        else:
            l = tint * (255 - l) + l  # retro-engineering of Excel behavior
        return PyxlWorkbook.hsl_to_rgb(t, s, l)

    def get_rgb_color(self, color):
        if color is None:
            return None
        if color.type == 'rgb':
            if color.rgb[0:2] == "00": # alpha = do not use color
                return None
            return self.tint_rgb(color.rgb[2:], color.tint)

        if color.type =='theme':
            try:
                rgb = self.theme[color.index]
                return self.tint_rgb(rgb, color.tint)
            except IndexError:
                raise PyxlSqlInternalError("Excel theme out of range ")

        raise PyxlSqlInternalError("Excel unknown color not theme nor rgb")


    @staticmethod
    def escape(excel_str: str):
        html_str = excel_str.encode('ascii', 'xmlcharrefreplace')
        html_str =  html_str.decode('utf-8')
        return html_str

    def get_html_style_from_cell(self, cell):
        # Manage bold/italic, center/left/right, number formats, date, formula

        cell_format = ""
        cell_style = 'style="'
        v_align = cell.alignment.vertical
        if v_align is  None:
            # excel defaults to bottom
            # html defaults to top
            cell_style += " vertical-align: bottom; "
        elif v_align == 'center':
            cell_style += " vertical-align: middle; "
        elif v_align == 'top':
            cell_style += " vertical-align: top; "

        cell_color = self.get_rgb_color(cell.fill.fgColor)
        if cell_color is not None:
            cell_format += f" bgcolor='#{cell_color}' "  # noqa

        h_align = cell.alignment.horizontal
        if h_align is not None:
            cell_style += f" text-align: {h_align}; "


        cell_font = cell.font
        if cell_font is not None :
            # rough font-size computation, from https://www.w3schools.com/css/css_font_size.asp
            # default html font-size = 16px = 1em
            # default excel font-size = 11
            # rough-estimate : html-font-size = excel*16/11 px
            cell_style += f"font-family:'{cell_font.name}';font-size:{int(cell_font.size*16/11)}px;"
            font_color = self.get_rgb_color(cell_font.color)
            if font_color is not None:
                cell_style += f"color:#{font_color};"
                # TODO: use mathplotlib to build a pie chart   # noqa
                # https://matplotlib.org/3.1.1/gallery/pie_and_polar_charts/pie_features
                # voir https://groups.google.com/g/openpyxl-users/c/v2FDsbDDTqU/m/rQWLAXZFkeUJ
            if cell_font.b:
                cell_style += "font-weight: bold; "
            if cell_font.i:
                cell_style += 'font-style: italic; '
            if cell_font.u:
                cell_style += 'text-decoration: underline; '

        if cell_style != "":
           cell_style += '"'

        return cell_format + cell_style

    def get_html_style_by_nb(self, sheet, line_nb: int, col_nb:int):
        cell = sheet.cell(line_nb, col_nb)
        return self.get_html_style_from_cell(cell)

    @staticmethod
    def get_html_value_by_nb(sheet, line_nb: int, col_nb: int):
        cell = sheet.cell(line_nb, col_nb)
        return PyxlSheet.get_html_value_from_cell(cell)

    def export_html(self, file_name: str, css: str=""):

        outs = self.local_open(file_name, "w")

        outs.writelines('<!DOCTYPE html>\n')
        outs.writelines('<html>\n')
        outs.writelines('<head>\n')
        outs.writelines('    <meta http-equiv=Content-Type content="text/html; charset=windows-1252">\n')
        outs.writelines(f'    <title>{self.filename}</title>\n')
        if css == "":
            outs.writelines('   <style>\n')
            outs.writelines('     html, body, h1, h2, h3, p { margin: 5; padding: 5; border: 0; \n')
            outs.writelines('            font-size: 100%; font-family: inherit; vertical-align: baseline; }\n')
            outs.writelines('     table, td, th { margin: 0; padding: 1; border: 1px solid black;\n')
            outs.writelines('            border-collapse: collapse; vertical-align: top;}\n')
            outs.writelines('     th  {text-align: left;}\n')

            outs.writelines('     H1 {background: #f8f8f8; width: 100%; border-bottom: 1px solid #ccc;  \n')
            outs.writelines('            border-top: 1px solid #ccc; font-size: 1.25em; display: inline-block;}\n')
            outs.writelines('     H2 {background: #888888; width: 100%; margin-top: .5em; font-size: 1em; display: inline-block;}\n')
            outs.writelines('   </style>\n')
        else:
            outs.writelines(f'    <link rel="stylesheet" href="{css}" type="text/css">\n')


        outs.writelines('</head>\n')
        outs.writelines('<body>\n')
        outs.writelines(f'    <h1>{self.filename}</h1>\n')

        for cur_sheet in self.wb.worksheets:
            outs.writelines('   <div  class="sheet">\n')
            outs.writelines(f'      <A id="{cur_sheet.title}"><h2>Sheet [{cur_sheet.title}]</h2></a>\n')
            outs.writelines('      <br><br>\n')
            outs.writelines('      <table>\n')
            outs.writelines('        <thead>\n')                                           # noqa
            outs.writelines('          <tr><th>1</th>\n')
            for col_nb in range(1, cur_sheet.max_column+1):
                col_name = cur_sheet.cell(1,col_nb).value or ''
                outs.write(f'            <th {self.get_html_style_by_nb(cur_sheet,1, col_nb)}>{col_name}</th>\n')
            outs.write('          </tr>\n')
            outs.write('        </thead>\n')                                               # noqa

            for lineno in range(2, cur_sheet.max_row+1):
                outs.write(f'         <tr><td>{lineno}</td>\n')
                for col_nb in range(1, cur_sheet.max_column+1):
                    outs.write(f'           <td {self.get_html_style_by_nb(cur_sheet,lineno, col_nb)}>')
                    val = self.get_html_value_by_nb(cur_sheet,lineno, col_nb)  or ''
                    outs.write(f'               {self.escape(val)}</td>\n')
                outs.write('         </tr>\n')
            outs.write('      </table>\n')
            outs.write('   </div>\n')
        outs.write('</body>\n')
        outs.write('</html>\n')

    def get_workbook(self, name):
        if name not in self.wbs:
            self.wbs[name] = PyxlWorkbook(self.file_path+'/'+name)
        return self.wbs[name]

    def import_module(self, module, sub_modules=None):
        mod = importlib.import_module(module)
        if sub_modules is None:
            self.imported[module] = mod
            return

        if sub_modules == '*':
            if hasattr(mod, '__all__'):
                item_list = mod.__all__
            else:
                raise PyxlSqlError(f"ERROR: IMPORT {module} SUBS {sub_modules} : ABORTING the import",
                                   "        The module does not contain a __all__ symbol that allows importing *")
        else:
            item_list = sub_modules.split()

        for item in item_list:
            self.imported[item] = getattr(mod, item)
